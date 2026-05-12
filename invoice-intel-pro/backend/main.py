import os
import json
import re
import ollama
import torch
import shutil
import time

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pathlib import Path
from typing import List

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Docling
from docling.document_converter import (
    DocumentConverter,
    PdfFormatOption,
    ImageFormatOption
)

from docling.datamodel.pipeline_options import (
    PdfPipelineOptions,
    RapidOcrOptions,
    AcceleratorOptions
)

from docling.datamodel.base_models import InputFormat


# =========================================================
# FASTAPI INITIALIZATION
# =========================================================

app = FastAPI()


# =========================================================
# CORS CONFIGURATION
# =========================================================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================================================
# DIRECTORY SETUP
# =========================================================

BASE_DIR = Path(__file__).resolve().parent

UPLOAD_DIR = BASE_DIR / "invoices"
OUTPUT_DIR = BASE_DIR / "output"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


# =========================================================
# MODEL CONFIGURATION
# =========================================================

MODEL_NAME = "gemma4:31b-cloud"


# =========================================================
# DEVICE CONFIGURATION
# =========================================================

compute_device = (
    "cuda"
    if torch.cuda.is_available()
    else (
        "mps"
        if torch.backends.mps.is_available()
        else "cpu"
    )
)

print(f"Using device: {compute_device}")
print(f"Using model: {MODEL_NAME}")


# =========================================================
# DOCLING CONFIGURATION
# =========================================================

pipeline_opts = PdfPipelineOptions()

pipeline_opts.do_ocr = True
pipeline_opts.ocr_options = RapidOcrOptions()

pipeline_opts.do_table_structure = True

pipeline_opts.accelerator_options = AcceleratorOptions(
    num_threads=os.cpu_count() or 4,
    device=compute_device
)

converter = DocumentConverter(
    format_options={
        InputFormat.PDF: PdfFormatOption(
            pipeline_options=pipeline_opts
        ),

        InputFormat.IMAGE: ImageFormatOption(
            pipeline_options=pipeline_opts
        )
    }
)


# =========================================================
# UTILITIES
# =========================================================

def clean_json_response(raw_text: str) -> dict:

    try:

        cleaned = re.sub(
            r"```json|```",
            "",
            raw_text
        ).strip()

        start = cleaned.find("{")
        end = cleaned.rfind("}") + 1

        if start != -1 and end != -1:
            cleaned = cleaned[start:end]

        return json.loads(cleaned)

    except Exception:
        return {}


def flatten_dict(d, parent_key='', sep=' '):

    items = []

    for k, v in d.items():

        new_key = (
            f"{parent_key}{sep}{k}"
            if parent_key
            else k
        )

        if isinstance(v, dict):

            items.extend(
                flatten_dict(
                    v,
                    new_key,
                    sep=sep
                ).items()
            )

        else:

            display_key = (
                str(new_key)
                .replace('_', ' ')
                .replace('.', ' ')
                .title()
            )

            items.append((display_key, v))

    return dict(items)


# =========================================================
# ROOT ENDPOINT
# =========================================================

@app.get("/")
async def root():

    return {
        "message": "AI Invoice Extractor Backend Running"
    }


# =========================================================
# MULTIPLE FILE EXTRACT ENDPOINT
# =========================================================

@app.post("/api/extract")
async def extract_invoice(
    files: List[UploadFile] = File(...)
):

    allowed_extensions = {
        ".pdf",
        ".png",
        ".jpg",
        ".jpeg",
        ".tiff",
        ".bmp"
    }

    # =====================================================
    # EXCEL SETUP
    # =====================================================

    wb = Workbook()

    ws = wb.active
    ws.title = "Invoice Extraction"

    target_columns = [
        "Source File",
        "Document Type",
        "Seller Name",
        "Seller GSTIN",
        "Invoice No",
        "Invoice Date",
        "Agency Address",
        "Email Address",
        "Mobile No",
        "Phone No",
        "Description",
        "HSN Code",
        "Quantity",
        "Unit Rate",
        "Taxable Amt",
        "CGST Amt",
        "SGST Amt",
        "IGST Amt",
        "Other Charges",
        "Round Off",
        "Grand Total",
        "Bank Name",
        "Bank A/C No",
        "IFSC Code"
    ]

    # =====================================================
    # HEADER STYLING
    # =====================================================

    for col_index, col_name in enumerate(
        target_columns,
        start=1
    ):

        cell = ws.cell(
            row=1,
            column=col_index,
            value=col_name
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            start_color="4F46E5",
            end_color="4F46E5",
            fill_type="solid"
        )

    current_excel_row = 2

    processed_files = []

    # =====================================================
    # PROCESS FILES ONE BY ONE
    # =====================================================

    for file in files:

        file_extension = Path(
            file.filename
        ).suffix.lower()

        if file_extension not in allowed_extensions:

            print(f"Skipping unsupported file: {file.filename}")
            continue

        file_path = UPLOAD_DIR / file.filename

        try:

            # =============================================
            # SAVE FILE
            # =============================================

            with open(file_path, "wb") as f:
                shutil.copyfileobj(file.file, f)

            print(f"\nProcessing File: {file.filename}")

            # =============================================
            # DOCLING CONVERSION
            # =============================================

            result = converter.convert(file_path)

            print("Document converted successfully")

            # =============================================
            # TABLE EXTRACTION
            # =============================================

            table_csv = ""

            if result.document.tables:

                table_csv = "\n".join([

                    table.export_to_dataframe(
                        doc=result.document
                    ).to_csv(index=False)

                    for table in result.document.tables

                ])

            # =============================================
            # MARKDOWN EXTRACTION
            # =============================================

            text_content = result.document.export_to_markdown()

            # =============================================
            # HEADER EXTRACTION
            # =============================================

            print("Extracting header information...")

            header_response = ollama.chat(
                model=MODEL_NAME,
                format="json",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Extract these invoice fields:\n\n"
                            "Document Type,\n"
                            "Seller Name,\n"
                            "Seller GSTIN,\n"
                            "Invoice No,\n"
                            "Invoice Date,\n"
                            "Agency Address,\n"
                            "Email Address,\n"
                            "Mobile No,\n"
                            "Phone No.\n\n"
                            "Return ONLY valid flat JSON."
                        )
                    },
                    {
                        "role": "user",
                        "content": text_content
                    }
                ]
            )

            header_data = clean_json_response(
                header_response["message"]["content"]
            )

            print("Header extraction completed")

            # =============================================
            # LINE ITEMS EXTRACTION
            # =============================================

            print("Extracting line items...")

            line_response = ollama.chat(
                model=MODEL_NAME,
                format="json",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Extract invoice line items.\n\n"
                            "Fields:\n"
                            "Description,\n"
                            "HSN Code,\n"
                            "Quantity,\n"
                            "Unit Rate,\n"
                            "Taxable Amt,\n"
                            "CGST Amt,\n"
                            "SGST Amt,\n"
                            "IGST Amt.\n\n"
                            "Return JSON with key 'items'."
                        )
                    },
                    {
                        "role": "user",
                        "content": (
                            table_csv
                            if table_csv.strip()
                            else text_content
                        )
                    }
                ]
            )

            line_data = clean_json_response(
                line_response["message"]["content"]
            )

            items = line_data.get("items", [])

            print("Line item extraction completed")

            # =============================================
            # FINANCIAL EXTRACTION
            # =============================================

            print("Extracting financial details...")

            financial_response = ollama.chat(
                model=MODEL_NAME,
                format="json",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Extract these financial fields:\n\n"
                            "Other Charges,\n"
                            "Round Off,\n"
                            "Grand Total,\n"
                            "Bank Name,\n"
                            "Bank A/C No,\n"
                            "IFSC Code.\n\n"
                            "Return ONLY valid flat JSON."
                        )
                    },
                    {
                        "role": "user",
                        "content": text_content
                    }
                ]
            )

            financial_data = clean_json_response(
                financial_response["message"]["content"]
            )

            print("Financial extraction completed")

            # =============================================
            # MERGE DATA
            # =============================================

            base_context = {
                "Source File": file.filename,
                **flatten_dict(header_data),
                **flatten_dict(financial_data)
            }

            if items:

                unified_rows = [
                    {
                        **base_context,
                        **flatten_dict(item)
                    }
                    for item in items
                ]

            else:

                unified_rows = [base_context]

            # =============================================
            # WRITE TO EXCEL
            # =============================================

            for row_dict in unified_rows:

                for col_index, col_name in enumerate(
                    target_columns,
                    start=1
                ):

                    target_clean = (
                        col_name.lower()
                        .replace(".", "")
                        .strip()
                    )

                    value = ""

                    for key, val in row_dict.items():

                        if target_clean in str(key).lower():

                            value = val
                            break

                    ws.cell(
                        row=current_excel_row,
                        column=col_index,
                        value=str(value)
                    )

                current_excel_row += 1

            processed_files.append(file.filename)

            print(f"{file.filename} processed successfully")

        except Exception as e:

            print(f"Error processing {file.filename}: {str(e)}")

        finally:

            if file_path.exists():
                os.remove(file_path)

    # =====================================================
    # AUTO COLUMN WIDTH
    # =====================================================

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value
            else 0
            for cell in column_cells
        )

        adjusted_width = min(length + 5, 50)

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = adjusted_width

    # =====================================================
    # SAVE EXCEL
    # =====================================================

    print("Generating final Excel file...")

    excel_filename = "final_invoice_extraction.xlsx"

    excel_path = OUTPUT_DIR / excel_filename

    wb.save(excel_path)

    time.sleep(1)

    print("Excel generated successfully")

    # =====================================================
    # RESPONSE
    # =====================================================

    return {
        "status": "success",
        "message": "All invoices processed successfully",
        "processed_files": processed_files,
        "total_files": len(processed_files),
        "excel_file": excel_filename
    }


# =========================================================
# DOWNLOAD ENDPOINT
# =========================================================

@app.get("/api/download/{filename}")
async def download_file(filename: str):

    excel_path = OUTPUT_DIR / filename

    if not excel_path.exists():

        raise HTTPException(
            status_code=404,
            detail="Excel file not found"
        )

    return FileResponse(
        path=excel_path,
        filename=filename,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# =========================================================
# MAIN
# =========================================================

if __name__ == "__main__":

    import uvicorn

    uvicorn.run(
        app,
        host="127.0.0.1",
        port=8000,
        reload=True
    )